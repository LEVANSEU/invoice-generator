import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
import re

st.title("Excel გენერატორი")

report_file = st.file_uploader("ატვირთე ანგარიშფაქტურების ფაილი (report.xlsx)", type=["xlsx"])
statement_file = st.file_uploader("ატვირთე საბანკო ამონაწერის ფაილი (statement.xlsx)", type=["xlsx"])

if report_file and statement_file:
    purchases_df = pd.read_excel(report_file, sheet_name='Grid')
    bank_df = pd.read_excel(statement_file)

    purchases_df['დასახელება'] = purchases_df['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
    purchases_df['საიდენტიფიკაციო კოდი'] = purchases_df['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])
    bank_df['P'] = bank_df.iloc[:, 15].astype(str).str.strip()
    bank_df['Amount'] = pd.to_numeric(bank_df.iloc[:, 3], errors='coerce').fillna(0)

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="ანგარიშფაქტურები კომპანიით")
    ws1.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურის №', 'ანგარიშფაქტურის თანხა', 'ჩარიცხული თანხა'])
    
    company_summaries = []  # Streamlit-სთვის info-ს შესანახად

    for company_id, group in purchases_df.groupby('საიდენტიფიკაციო კოდი'):
        company_name = group['დასახელება'].iloc[0]
        unique_invoices = group.groupby('სერია №')['ღირებულება დღგ და აქციზის ჩათვლით'].sum().reset_index()
        company_invoice_sum = unique_invoices['ღირებულება დღგ და აქციზის ჩათვლით'].sum()
        
        # ზუსტი row კომპანიისთვის
        company_summary_row = ws1.max_row + 1
        payment_formula = f"=SUMIF(საბანკოამონაწერი!P:P, B{company_summary_row}, საბანკოამონაწერი!D:D)"
        ws1.append([company_name, company_id, '', company_invoice_sum, payment_formula])
        
        for _, row in unique_invoices.iterrows():
            ws1.append(['', '', row['სერია №'], row['ღირებულება დღგ და აქციზის ჩათვლით'], ''])
        
        company_summaries.append((company_name, company_id, company_invoice_sum))

    ws2 = wb.create_sheet(title="დეტალური მონაცემები")
    ws2.append(purchases_df.columns.tolist())
    for row in purchases_df.itertuples(index=False):
        ws2.append(row)

    ws3 = wb.create_sheet(title="საბანკოამონაწერი")
    ws3.append(bank_df.columns.tolist())
    for row in bank_df.itertuples(index=False):
        ws3.append(row)

    ws4 = wb.create_sheet(title="ანგარიშფაქტურის დეტალები")
    invoice_details_df = purchases_df[['სერია №', 'საქონელი / მომსახურება', 'ზომის ერთეული', 'რაოდ.', 'ღირებულება დღგ და აქციზის ჩათვლით']].copy()
    invoice_details_df.rename(columns={'სერია №': 'ინვოისის №'}, inplace=True)
    ws4.append(invoice_details_df.columns.tolist())
    for row in invoice_details_df.itertuples(index=False):
        ws4.append(row)

    ws5 = wb.create_sheet(title="გადარიცხვები_უბმოლოდ")
    missing_payments = bank_df[~bank_df['P'].isin(purchases_df['საიდენტიფიკაციო კოდი'])]
    ws5.append(missing_payments.columns.tolist())
    for row in missing_payments.itertuples(index=False):
        ws5.append(row)

    ws6 = wb.create_sheet(title="განახლებული ამონაწერი")
    ws6.append(bank_df.columns.tolist())
    for row in bank_df.itertuples(index=False):
        ws6.append(row)

    ws7 = wb.create_sheet(title="კომპანიების_ჯამები")
    ws7.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურების ჯამი', 'ჩარიცხული თანხა'])
    for idx, (company_name, company_id, invoice_sum) in enumerate(company_summaries, start=2):
        payment_formula = f"=SUMIF(საბანკოამონაწერი!P:P, B{idx}, საბანკოამონაწერი!D:D)"
        ws7.append([company_name, company_id, invoice_sum, payment_formula])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # ✅ Streamlit კომპანიის ღილაკებით და სვეტებით
    st.subheader("📋 კომპანიების ჩამონათვალი")

    st.write("| დასახელება | საიდენტიფიკაციო კოდი | ინვოისების ჯამი | ჩარიცხვა | სხვაობა |")
    st.write("|------------|------------------------|------------------|-----------|----------|")

    for name, company_id, invoice_sum in company_summaries:
        col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1.5])

        with col1:
            if st.button(f"{name}", key=f"name_{company_id}"):
                st.session_state['selected_company'] = name
        with col2:
            if st.button(f"{company_id}", key=f"id_{company_id}"):
                st.session_state['selected_company'] = company_id

        paid_sum = bank_df[bank_df["P"] == str(company_id)]["Amount"].sum()
        difference = invoice_sum - paid_sum

        with col3:
            st.write(f"{invoice_sum:,.2f}")
        with col4:
            st.write(f"{paid_sum:,.2f}")
        with col5:
            st.write(f"{difference:,.2f}")

    if 'selected_company' in st.session_state:
        st.info(f"🔎 არჩეული კომპანია: **{st.session_state['selected_company']}**")

    st.success("✅ ფაილი მზადაა! ჩამოტვირთე აქედან:")
    st.download_button(
        label="⬇️ ჩამოტვირთე Excel ფაილი",
        data=output,
        file_name="საბოლოო_ფაილი.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
